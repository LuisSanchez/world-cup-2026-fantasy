"""Build admin Excel exports (leaderboard + per-player match scores)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import is_leaderboard_excluded_email
from app.models import Match, Prediction, User


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1B4332")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(length + 2, 10)


def build_scores_workbook(db: Session) -> BytesIO:
    """
    Two sheets:
      1. Leaderboard — rank, points, prediction count (excludes spectator super-admin).
      2. Puntuaciones — every player's points per match (incl. all users in DB).
    """
    wb = Workbook()

    # ── Sheet 1: Leaderboard ─────────────────────────────
    ws_lb = wb.active
    ws_lb.title = "Leaderboard"

    lb_headers = [
        "Posición",
        "Nombre",
        "Email",
        "Puntos totales",
        "Pronósticos",
        "Admin",
    ]
    ws_lb.append(lb_headers)
    _style_header(ws_lb)

    users_ranked = db.query(User).order_by(User.total_points.desc(), User.name).all()
    rank = 1
    for u in users_ranked:
        if is_leaderboard_excluded_email(u.email):
            continue
        pred_count = db.query(Prediction).filter(Prediction.user_id == u.id).count()
        ws_lb.append(
            [
                rank,
                u.name or u.email.split("@")[0],
                u.email,
                u.total_points or 0,
                pred_count,
                "Sí" if u.is_admin else "No",
            ]
        )
        rank += 1
    _autosize(ws_lb)
    ws_lb.freeze_panes = "A2"
    ws_lb.auto_filter.ref = ws_lb.dimensions

    # ── Sheet 2: Per-player current scores (match breakdown) ──
    ws_sc = wb.create_sheet("Puntuaciones")
    sc_headers = [
        "Jugador",
        "Email",
        "Puntos totales (jugador)",
        "Partido #",
        "Fase",
        "Local",
        "Visitante",
        "Resultado real",
        "Finalizado",
        "Pronóstico",
        "Pts goles",
        "Pts resultado",
        "Pts partido",
    ]
    ws_sc.append(sc_headers)
    _style_header(ws_sc)

    matches = db.query(Match).order_by(Match.match_number).all()
    all_users = db.query(User).order_by(User.name, User.email).all()

    for u in all_users:
        if is_leaderboard_excluded_email(u.email):
            continue
        preds = {
            p.match_id: p
            for p in db.query(Prediction).filter(Prediction.user_id == u.id).all()
        }
        display_name = u.name or u.email.split("@")[0]
        total = u.total_points or 0

        if not matches:
            ws_sc.append(
                [
                    display_name,
                    u.email,
                    total,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "(sin partidos)",
                    "",
                    "",
                    "",
                ]
            )
            continue

        for m in matches:
            p = preds.get(m.id)
            real = ""
            if m.home_score is not None and m.away_score is not None:
                real = f"{m.home_score}-{m.away_score}"
            pred_str = ""
            pg = pr = pt = ""
            if p:
                pred_str = f"{p.home_score}-{p.away_score}"
                pg = p.points_goals if p.points_goals is not None else 0
                pr = p.points_result if p.points_result is not None else 0
                pt = p.points_total if p.points_total is not None else 0
            else:
                pred_str = "—"

            ws_sc.append(
                [
                    display_name,
                    u.email,
                    total,
                    m.match_number,
                    m.stage or "",
                    m.home_team,
                    m.away_team,
                    real or "—",
                    "Sí" if m.is_finished else "No",
                    pred_str,
                    pg if pg != "" else "",
                    pr if pr != "" else "",
                    pt if pt != "" else "",
                ]
            )

    _autosize(ws_sc)
    ws_sc.freeze_panes = "A2"
    ws_sc.auto_filter.ref = ws_sc.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_filename() -> str:
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    return f"wc_fantasy_puntuaciones_{stamp}.xlsx"
