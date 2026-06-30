"""Excel export builder tests."""

from types import SimpleNamespace
from unittest.mock import MagicMock

from openpyxl import load_workbook

from app.export_excel import build_scores_workbook, export_filename


def test_export_filename_has_xlsx():
    assert export_filename().endswith(".xlsx")
    assert "wc_fantasy" in export_filename()


def test_build_scores_workbook_two_sheets():
    db = MagicMock()
    u1 = SimpleNamespace(
        id=1,
        email="a@x.com",
        name="Ana",
        total_points=10,
        is_admin=False,
    )
    u_spectator = SimpleNamespace(
        id=2,
        email="admin@localhost.dev",
        name="Spectator",
        total_points=0,
        is_admin=True,
    )
    m1 = SimpleNamespace(
        id=1,
        match_number=1,
        stage="group",
        home_team="México",
        away_team="Sudáfrica",
        home_score=1,
        away_score=0,
        is_finished=True,
    )
    pred = SimpleNamespace(
        match_id=1,
        home_score=1,
        away_score=0,
        points_goals=2,
        points_result=1,
        points_total=3,
    )

    ranked_q = MagicMock()
    ranked_q.order_by.return_value.all.return_value = [u1, u_spectator]

    all_users_q = MagicMock()
    all_users_q.order_by.return_value.all.return_value = [u1, u_spectator]

    matches_q = MagicMock()
    matches_q.order_by.return_value.all.return_value = [m1]

    pred_count_q = MagicMock()
    pred_count_q.filter.return_value.count.return_value = 1

    pred_list_q = MagicMock()
    pred_list_q.filter.return_value.all.return_value = [pred]

    db.query.side_effect = [
        ranked_q,  # users ranked
        pred_count_q,  # count for u1
        matches_q,  # matches
        all_users_q,  # all users
        pred_list_q,  # preds for u1
    ]

    buf = build_scores_workbook(db)
    wb = load_workbook(buf)
    assert "Leaderboard" in wb.sheetnames
    assert "Puntuaciones" in wb.sheetnames
    lb = wb["Leaderboard"]
    assert lb["A1"].value == "Posición"
    assert lb["B2"].value == "Ana"
    sc = wb["Puntuaciones"]
    assert sc["A1"].value == "Jugador"
    assert sc["A2"].value == "Ana"
    assert sc["M2"].value == 3  # pts partido
